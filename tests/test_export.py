"""Tests for Export Functionality (v0.1.2 backfill + v0.1.3)."""
import pytest
import tempfile
import os
from pathlib import Path
from primerlab.core.output import OutputManager
from primerlab.core.models import WorkflowResult, Primer, Amplicon
from primerlab.core.models.metadata import RunMetadata
from primerlab.core.models.qc import QCResult


class TestOutputManager:
    """Tests for OutputManager class."""
    
    @pytest.fixture
    def sample_result(self):
        """Create a sample WorkflowResult for testing."""
        primers = {
            "forward": Primer(
                id="forward",
                sequence="ATGCGATCGATCGATCGC",
                tm=60.0,
                gc=55.0,
                length=18,
                start=10,
                end=28
            ),
            "reverse": Primer(
                id="reverse",
                sequence="GCTAGCTAGCTAGCTAG",
                tm=59.0,
                gc=52.0,
                length=17,
                start=100,
                end=117
            )
        }
        
        amplicons = [
            Amplicon(
                start=10,
                end=117,
                length=107,
                sequence="ATGC" * 25,
                gc=50.0,
                tm_forward=60.0,
                tm_reverse=59.0
            )
        ]
        
        metadata = RunMetadata(
            workflow="pcr",
            timestamp="2024-01-01T00:00:00Z",
            version="0.1.3-dev",
            parameters={}
        )
        
        qc = QCResult(
            hairpin_ok=True,
            homodimer_ok=True,
            heterodimer_ok=True,
            tm_balance_ok=True,
            hairpin_dg=0.0,
            homodimer_dg=0.0,
            heterodimer_dg=0.0,
            tm_diff=1.0,
            warnings=[],
            errors=[]
        )
        
        return WorkflowResult(
            workflow="pcr",
            primers=primers,
            amplicons=amplicons,
            metadata=metadata,
            qc=qc,
            raw={}
        )
    
    def test_output_manager_creates_directory(self):
        """OutputManager should create output directory."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = os.path.join(tmpdir, "test_output")
            mgr = OutputManager(output_dir, "pcr")
            
            assert mgr.run_dir.exists()
            assert "PCR" in str(mgr.run_dir)
    
    def test_save_json(self, sample_result):
        """save_json should create valid JSON file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            mgr = OutputManager(tmpdir, "pcr")
            mgr.save_json(sample_result)
            
            json_path = mgr.run_dir / "result.json"
            assert json_path.exists()
    
    def test_save_csv(self, sample_result):
        """save_csv should create CSV file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            mgr = OutputManager(tmpdir, "pcr")
            mgr.save_csv(sample_result)
            
            csv_path = mgr.run_dir / "primers.csv"
            assert csv_path.exists()
            
            # Check content
            content = csv_path.read_text()
            assert "forward" in content
            assert "reverse" in content
    
    def test_save_ordering_format_idt(self, sample_result):
        """save_ordering_format should create IDT format file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            mgr = OutputManager(tmpdir, "pcr")
            mgr.save_ordering_format(sample_result, "idt")
            
            idt_path = mgr.run_dir / "order_idt.csv"
            assert idt_path.exists()
    
    def test_save_ordering_format_sigma(self, sample_result):
        """save_ordering_format should create Sigma format file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            mgr = OutputManager(tmpdir, "pcr")
            mgr.save_ordering_format(sample_result, "sigma")
            
            sigma_path = mgr.run_dir / "order_sigma.csv"
            assert sigma_path.exists()
    
    def test_save_ordering_format_thermo(self, sample_result):
        """save_ordering_format should create Thermo format file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            mgr = OutputManager(tmpdir, "pcr")
            mgr.save_ordering_format(sample_result, "thermo")
            
            thermo_path = mgr.run_dir / "order_thermo.csv"
            assert thermo_path.exists()


class TestExportFormats:
    """Tests for vendor-specific export formats."""
    
    @pytest.fixture
    def sample_result(self):
        """Create a sample result."""
        primers = {
            "forward": Primer(
                id="forward",
                sequence="ATGCGATCGATCGATCGC",
                tm=60.0,
                gc=55.0,
                length=18,
                start=10,
                end=28
            )
        }
        
        metadata = RunMetadata(
            workflow="pcr",
            timestamp="2024-01-01T00:00:00Z",
            version="0.1.3-dev"
        )
        
        qc = QCResult(
            hairpin_ok=True,
            homodimer_ok=True,
            heterodimer_ok=True,
            tm_balance_ok=True,
            hairpin_dg=0.0,
            homodimer_dg=0.0,
            heterodimer_dg=0.0,
            tm_diff=0.0
        )
        
        return WorkflowResult(
            workflow="pcr",
            primers=primers,
            amplicons=[],
            metadata=metadata,
            qc=qc,
            raw={}
        )
    
    def test_all_formats_have_sequence(self, sample_result):
        """All export formats should include primer sequences."""
        with tempfile.TemporaryDirectory() as tmpdir:
            mgr = OutputManager(tmpdir, "pcr")
            
            for fmt in ["idt", "sigma", "thermo"]:
                mgr.save_ordering_format(sample_result, fmt)
                
                file_path = mgr.run_dir / f"order_{fmt}.csv"
                content = file_path.read_text()
                
                # Should contain the sequence
                assert "ATGCGATCGATCGATCGC" in content


class TestExcelExport:
    """Tests for Excel export functionality (v0.1.4)."""
    
    @pytest.fixture
    def sample_result(self):
        """Create a sample WorkflowResult for testing."""
        primers = {
            "forward": Primer(
                id="GAPDH_F1",
                sequence="CCCACTCCTCCACCTTTGAC",
                tm=60.0,
                gc=60.0,
                length=20,
                start=29,
                end=48,
                hairpin_dg=0.0,
                homodimer_dg=-2.5,
            ),
            "reverse": Primer(
                id="GAPDH_R1",
                sequence="TCCTCTTGTGCTCTTGCTGG",
                tm=60.0,
                gc=55.0,
                length=20,
                start=188,
                end=207,
                hairpin_dg=-1.1,
                homodimer_dg=-4.1,
            ),
        }
        
        metadata = RunMetadata(
            workflow="pcr",
            timestamp="2025-12-17T12:00:00+00:00",
            version="0.1.4",
        )
        
        qc = QCResult(
            hairpin_ok=True,
            homodimer_ok=True,
            heterodimer_ok=True,
            tm_balance_ok=True,
            hairpin_dg=-1.1,
            homodimer_dg=-4.1,
            heterodimer_dg=-3.2,
            tm_diff=0.0,
            quality_score=100,
            quality_category="Excellent",
            quality_category_emoji="✅",
        )
        
        return WorkflowResult(
            workflow="pcr",
            primers=primers,
            amplicons=[],
            metadata=metadata,
            qc=qc,
        )
    
    def test_excel_export_creates_file(self, sample_result):
        """Excel export should create a .xlsx file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_mgr = OutputManager(tmpdir, "pcr")
            output_mgr.save_excel(sample_result)
            
            xlsx_path = output_mgr.run_dir / "primers.xlsx"
            assert xlsx_path.exists(), "Excel file should be created"
    
    def test_excel_export_has_correct_sheets(self, sample_result):
        """Excel file should have Primers and QC Summary sheets."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_mgr = OutputManager(tmpdir, "pcr")
            output_mgr.save_excel(sample_result)
            
            xlsx_path = output_mgr.run_dir / "primers.xlsx"
            wb = load_workbook(xlsx_path)
            
            assert "Primers" in wb.sheetnames, "Should have Primers sheet"
            assert "QC Summary" in wb.sheetnames, "Should have QC Summary sheet"
    
    def test_excel_export_primer_data(self, sample_result):
        """Excel should contain correct primer data."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_mgr = OutputManager(tmpdir, "pcr")
            output_mgr.save_excel(sample_result)
            
            xlsx_path = output_mgr.run_dir / "primers.xlsx"
            wb = load_workbook(xlsx_path)
            ws = wb["Primers"]
            
            # Check header row
            assert ws.cell(1, 1).value == "Name"
            assert ws.cell(1, 2).value == "Sequence"
            
            # Check data row contains primer ID
            data_values = [ws.cell(2, col).value for col in range(1, 10)]
            assert "GAPDH_F1" in data_values, "Should contain forward primer ID"


class TestHTMLExport:
    """Tests for HTML report export functionality (v0.1.4)."""
    
    @pytest.fixture
    def sample_result(self):
        """Create a sample WorkflowResult for testing."""
        primers = {
            "forward": Primer(
                id="GAPDH_F1",
                sequence="CCCACTCCTCCACCTTTGAC",
                tm=60.0,
                gc=60.0,
                length=20,
                hairpin_dg=0.0,
            ),
        }
        
        metadata = RunMetadata(
            workflow="pcr",
            timestamp="2025-12-17T12:00:00+00:00",
            version="0.1.4",
        )
        
        qc = QCResult(
            hairpin_ok=True,
            homodimer_ok=True,
            heterodimer_ok=True,
            tm_balance_ok=True,
            hairpin_dg=0.0,
            homodimer_dg=0.0,
            heterodimer_dg=0.0,
            tm_diff=0.0,
            quality_score=100,
            quality_category="Excellent",
        )
        
        return WorkflowResult(
            workflow="pcr",
            primers=primers,
            amplicons=[],
            metadata=metadata,
            qc=qc,
        )
    
    def test_html_export_creates_file(self, sample_result):
        """HTML export should create an .html file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_mgr = OutputManager(tmpdir, "pcr")
            output_mgr.save_html(sample_result)
            
            html_path = output_mgr.run_dir / "report.html"
            assert html_path.exists(), "HTML file should be created"
    
    def test_html_contains_primer_info(self, sample_result):
        """HTML should contain primer sequences."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_mgr = OutputManager(tmpdir, "pcr")
            output_mgr.save_html(sample_result)
            
            html_path = output_mgr.run_dir / "report.html"
            content = html_path.read_text()
            
            assert "GAPDH_F1" in content, "Should contain primer ID"
            assert "CCCACTCCTCCACCTTTGAC" in content, "Should contain primer sequence"
    
    def test_html_contains_quality_score(self, sample_result):
        """HTML should display quality score."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_mgr = OutputManager(tmpdir, "pcr")
            output_mgr.save_html(sample_result)
            
            html_path = output_mgr.run_dir / "report.html"
            content = html_path.read_text()
            
            assert "100" in content, "Should contain quality score"
            assert "Excellent" in content, "Should contain quality category"
    
    def test_html_is_self_contained(self, sample_result):
        """HTML should have embedded CSS (no external dependencies)."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_mgr = OutputManager(tmpdir, "pcr")
            output_mgr.save_html(sample_result)
            
            html_path = output_mgr.run_dir / "report.html"
            content = html_path.read_text()
            
            assert "<style>" in content, "Should have embedded CSS"
            assert "</style>" in content, "Should close style tag"


class TestIDTBulkExport:
    """Tests for IDT Bulk Order export functionality (v0.1.4)."""
    
    @pytest.fixture
    def sample_result(self):
        """Create a sample WorkflowResult for testing."""
        primers = {
            "forward": Primer(
                id="GAPDH_F1",
                sequence="CCCACTCCTCCACCTTTGAC",
                tm=60.0,
                gc=60.0,
                length=20,
            ),
            "reverse": Primer(
                id="GAPDH_R1",
                sequence="TCCTCTTGTGCTCTTGCTGG",
                tm=60.0,
                gc=55.0,
                length=20,
            ),
        }
        
        metadata = RunMetadata(
            workflow="pcr",
            timestamp="2025-12-17T12:00:00+00:00",
            version="0.1.4",
        )
        
        qc = QCResult(
            hairpin_ok=True,
            homodimer_ok=True,
            heterodimer_ok=True,
            tm_balance_ok=True,
            hairpin_dg=0.0,
            homodimer_dg=0.0,
            heterodimer_dg=0.0,
            tm_diff=0.0,
        )
        
        return WorkflowResult(
            workflow="pcr",
            primers=primers,
            amplicons=[],
            metadata=metadata,
            qc=qc,
        )
    
    def test_idt_bulk_creates_file(self, sample_result):
        """IDT bulk export should create a file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_mgr = OutputManager(tmpdir, "pcr")
            output_mgr.save_idt_bulk_order(sample_result)
            
            idt_path = output_mgr.run_dir / "idt_bulk_order.xlsx"
            assert idt_path.exists(), "IDT bulk order file should be created"
    
    def test_idt_bulk_has_well_positions(self, sample_result):
        """IDT bulk order should have well positions (A1, A2, etc.)."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_mgr = OutputManager(tmpdir, "pcr")
            output_mgr.save_idt_bulk_order(sample_result)
            
            idt_path = output_mgr.run_dir / "idt_bulk_order.xlsx"
            wb = load_workbook(idt_path)
            ws = wb.active
            
            # Check for well positions in data
            headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
            
            # Find Well Position column
            well_col = None
            for idx, h in enumerate(headers, 1):
                if h and "Well" in str(h):
                    well_col = idx
                    break
            
            assert well_col is not None, "Should have Well Position column"
            
            # Check data has A1, A2 etc
            positions = [ws.cell(row, well_col).value for row in range(2, ws.max_row + 1)]
            assert any(pos and pos.startswith("A") for pos in positions), "Should have well positions like A1, A2"
    
    def test_idt_format_has_required_columns(self, sample_result):
        """IDT format should have Name, Sequence columns at minimum."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_mgr = OutputManager(tmpdir, "pcr")
            output_mgr.save_idt_bulk_order(sample_result)
            
            idt_path = output_mgr.run_dir / "idt_bulk_order.xlsx"
            wb = load_workbook(idt_path)
            ws = wb.active
            
            headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
            
            # IDT requires at least Name and Sequence
            assert any("Name" in str(h) for h in headers if h), "Should have Name column"
            assert any("Sequence" in str(h) for h in headers if h), "Should have Sequence column"

