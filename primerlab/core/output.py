import json
import csv
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List
from primerlab.core.models import WorkflowResult
from primerlab.core.logger import get_logger
from primerlab.core.exceptions import PrimerLabException

logger = get_logger()

class OutputManager:
    """
    Handles creation of run directories and saving of results.
    """
    def __init__(self, base_dir: str, workflow_name: str):
        self.base_dir = Path(base_dir)
        self.workflow_name = workflow_name
        self.run_dir = self._create_run_dir()

    def _create_run_dir(self) -> Path:
        """Creates a timestamped run directory."""
        # Format: YYYYMMDD_HHMMSS_WORKFLOW (e.g., 20251127_115143_PCR)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        run_name = f"{timestamp}_{self.workflow_name.upper()}"
        full_path = self.base_dir / run_name

        try:
            full_path.mkdir(parents=True, exist_ok=True)
            logger.info(f"Output directory created: {full_path}")
            return full_path
        except Exception as e:
            raise PrimerLabException(f"Failed to create output directory: {e}")

    def save_json(self, result: WorkflowResult, filename: str = "result.json"):
        """Saves WorkflowResult to a JSON file."""
        file_path = self.run_dir / filename
        try:
            data = result.to_dict()
            with open(file_path, 'w') as f:
                json.dump(data, f, indent=2)
            logger.info(f"Result saved to: {file_path}")
        except Exception as e:
            logger.error(f"Failed to save JSON: {e}")

    def save_csv(self, result: WorkflowResult, filename: str = "primers.csv"):
        """Saves primers to a CSV file for easy spreadsheet viewing."""
        file_path = self.run_dir / filename
        try:
            primers = result.primers
            if not primers:
                logger.warning("No primers to export to CSV.")
                return

            # Define CSV columns
            fieldnames = [
                "Name", "Sequence", "Length", "Tm", "GC%", 
                "Hairpin_dG", "Homodimer_dG", "Heterodimer_dG",
                "Start", "End", "Warnings"
            ]

            with open(file_path, 'w', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()

                for name, primer in primers.items():
                    writer.writerow({
                        "Name": primer.id,
                        "Sequence": primer.sequence,
                        "Length": primer.length,
                        "Tm": round(primer.tm, 2),
                        "GC%": round(primer.gc, 2),
                        "Hairpin_dG": round(primer.hairpin_dg, 2) if primer.hairpin_dg else "",
                        "Homodimer_dG": round(primer.homodimer_dg, 2) if primer.homodimer_dg else "",
                        "Heterodimer_dG": round(primer.heterodimer_dg, 2) if primer.heterodimer_dg else "",
                        "Start": primer.start if primer.start else "",
                        "End": primer.end if primer.end else "",
                        "Warnings": "; ".join(primer.warnings) if primer.warnings else ""
                    })

            logger.info(f"CSV export saved to: {file_path}")
        except Exception as e:
            logger.error(f"Failed to save CSV: {e}")

    def save_ordering_format(self, result: WorkflowResult, vendor: str = "idt"):
        """
        Saves primers in vendor-specific ordering format.
        Supported vendors: idt, sigma, thermo
        """
        primers = result.primers
        if not primers:
            logger.warning("No primers to export for ordering.")
            return

        vendor = vendor.lower()
        filename = f"order_{vendor}.csv"
        file_path = self.run_dir / filename

        try:
            with open(file_path, 'w', newline='') as f:
                if vendor == "idt":
                    # IDT format: Name, Sequence, Scale, Purification
                    writer = csv.writer(f)
                    writer.writerow(["Name", "Sequence", "Scale", "Purification"])
                    for name, primer in primers.items():
                        writer.writerow([primer.id, primer.sequence, "25nm", "STD"])

                elif vendor == "sigma":
                    # Sigma format: Oligo Name, Sequence (5' to 3')
                    writer = csv.writer(f)
                    writer.writerow(["Oligo Name", "Sequence (5' to 3')"])
                    for name, primer in primers.items():
                        writer.writerow([primer.id, primer.sequence])

                elif vendor == "thermo":
                    # Thermo Fisher format: Name, Sequence, Scale
                    writer = csv.writer(f)
                    writer.writerow(["Name", "Sequence", "Scale"])
                    for name, primer in primers.items():
                        writer.writerow([primer.id, primer.sequence, "25 nmole"])

                else:
                    logger.warning(f"Unknown vendor '{vendor}'. Using generic format.")
                    writer = csv.writer(f)
                    writer.writerow(["Name", "Sequence"])
                    for name, primer in primers.items():
                        writer.writerow([primer.id, primer.sequence])

            logger.info(f"Ordering format ({vendor.upper()}) saved to: {file_path}")
        except Exception as e:
            logger.error(f"Failed to save ordering format: {e}")

    def save_debug_data(self, data: Dict[str, Any], filename: str = "debug_raw.json"):
        """Saves raw debug data."""
        debug_dir = self.run_dir / "debug"
        debug_dir.mkdir(exist_ok=True)

        file_path = debug_dir / filename
        try:
            with open(file_path, 'w') as f:
                # Handle non-serializable objects if necessary, but for now assume dict is clean
                json.dump(data, f, indent=2, default=str)
            logger.debug(f"Debug data saved to: {file_path}")
        except Exception as e:
            logger.warning(f"Failed to save debug data: {e}")

    def save_excel(self, result: WorkflowResult, filename: str = "primers.xlsx"):
        """
        Saves primers to an Excel file with formatting and color coding.
        Requires openpyxl.
        
        v0.1.4: New export format
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not installed. Run: pip install openpyxl")
            return

        primers = result.primers
        if not primers:
            logger.warning("No primers to export to Excel.")
            return

        file_path = self.run_dir / filename

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Primers"

            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers
            headers = ["Name", "Sequence", "Length", "Tm (°C)", "GC%", 
                      "Hairpin ΔG", "Homodimer ΔG", "Start", "End"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Data rows
            row = 2
            for name, primer in primers.items():
                ws.cell(row=row, column=1, value=primer.id).border = thin_border
                ws.cell(row=row, column=2, value=primer.sequence).border = thin_border
                ws.cell(row=row, column=3, value=primer.length).border = thin_border

                tm_cell = ws.cell(row=row, column=4, value=round(primer.tm, 2))
                tm_cell.border = thin_border
                if 58 <= primer.tm <= 62:
                    tm_cell.fill = good_fill

                gc_cell = ws.cell(row=row, column=5, value=round(primer.gc, 2))
                gc_cell.border = thin_border
                if 40 <= primer.gc <= 60:
                    gc_cell.fill = good_fill
                elif primer.gc < 35 or primer.gc > 65:
                    gc_cell.fill = warn_fill

                ws.cell(row=row, column=6, value=round(primer.hairpin_dg, 2) if primer.hairpin_dg else "").border = thin_border
                ws.cell(row=row, column=7, value=round(primer.homodimer_dg, 2) if primer.homodimer_dg else "").border = thin_border
                ws.cell(row=row, column=8, value=primer.start if primer.start else "").border = thin_border
                ws.cell(row=row, column=9, value=primer.end if primer.end else "").border = thin_border

                row += 1

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            ws.column_dimensions['B'].width = 30  # Sequence column wider

            # QC Summary Sheet
            if result.qc:
                ws_qc = wb.create_sheet("QC Summary")
                qc_headers = ["Metric", "Value", "Status"]
                for col, header in enumerate(qc_headers, 1):
                    cell = ws_qc.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill

                qc_data = [
                    ("Tm Difference", f"{result.qc.tm_diff:.2f}°C", "✅" if result.qc.tm_balance_ok else "⚠️"),
                    ("Hairpin ΔG", f"{result.qc.hairpin_dg:.2f}", "✅" if result.qc.hairpin_ok else "⚠️"),
                    ("Homodimer ΔG", f"{result.qc.homodimer_dg:.2f}", "✅" if result.qc.homodimer_ok else "⚠️"),
                ]

                if result.qc.quality_score is not None:
                    qc_data.insert(0, ("Quality Score", f"{result.qc.quality_score}/100", result.qc.quality_category or ""))

                for row_idx, (metric, value, status) in enumerate(qc_data, 2):
                    ws_qc.cell(row=row_idx, column=1, value=metric)
                    ws_qc.cell(row=row_idx, column=2, value=value)
                    ws_qc.cell(row=row_idx, column=3, value=status)

            wb.save(file_path)
            logger.info(f"Excel export saved to: {file_path}")

        except Exception as e:
            logger.error(f"Failed to save Excel: {e}")

    def save_idt_bulk_order(self, result: WorkflowResult, filename: str = "idt_bulk_order.xlsx"):
        """
        Saves primers in IDT bulk upload format with plate position.
        
        v0.1.4: New export format
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.warning("openpyxl not installed. Run: pip install openpyxl")
            return

        primers = result.primers
        if not primers:
            logger.warning("No primers to export for IDT bulk order.")
            return

        file_path = self.run_dir / filename

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "IDT Bulk Order"

            # IDT format headers
            headers = ["Well Position", "Name", "Sequence", "Scale", "Purification"]
            header_font = Font(bold=True)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font

            # Assign well positions (A1, A2, B1, B2, etc.)
            wells = ["A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8", "A9", "A10", "A11", "A12"]
            row = 2
            well_idx = 0

            for name, primer in primers.items():
                well = wells[well_idx] if well_idx < len(wells) else f"A{well_idx + 1}"

                ws.cell(row=row, column=1, value=well)
                ws.cell(row=row, column=2, value=primer.id)
                ws.cell(row=row, column=3, value=primer.sequence)
                ws.cell(row=row, column=4, value="25nm")
                ws.cell(row=row, column=5, value="STD")

                row += 1
                well_idx += 1

            wb.save(file_path)
            logger.info(f"IDT bulk order saved to: {file_path}")

        except Exception as e:
            logger.error(f"Failed to save IDT bulk order: {e}")

    def save_html(self, result: WorkflowResult, filename: str = "report.html"):
        """
        Saves a standalone HTML report.
        
        v0.1.4: New export format
        """
        from primerlab.core.html_report import generate_html_report

        file_path = self.run_dir / filename

        try:
            html_content = generate_html_report(result)

            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(html_content)

            logger.info(f"HTML report saved to: {file_path}")

        except Exception as e:
            logger.error(f"Failed to save HTML report: {e}")

    def save_benchling_csv(self, result: WorkflowResult, filename: str = "benchling_primers.csv"):
        """
        Saves primers in Benchling-compatible CSV format.
        
        Benchling import format:
        - Name: Primer name
        - Bases: DNA sequence
        - Notes: Optional notes (Tm, GC%, etc.)
        
        v0.1.5: New export format
        """
        primers = result.primers
        if not primers:
            logger.warning("No primers to export for Benchling.")
            return

        file_path = self.run_dir / filename

        try:
            with open(file_path, 'w', newline='') as f:
                writer = csv.writer(f)
                # Benchling standard columns
                writer.writerow(["Name", "Bases", "Notes"])

                for name, primer in primers.items():
                    # Build notes with key metrics
                    notes = f"Tm={primer.tm:.1f}°C, GC={primer.gc:.1f}%"
                    if primer.length:
                        notes += f", {primer.length}nt"

                    writer.writerow([
                        primer.id,
                        primer.sequence,
                        notes
                    ])

            logger.info(f"Benchling export saved to: {file_path}")
        except Exception as e:
            logger.error(f"Failed to save Benchling CSV: {e}")
