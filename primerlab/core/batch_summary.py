"""
PrimerLab Batch Summary Module

v0.1.5: Generates consolidated summary reports for batch runs.

This module follows the architecture rules:
- Core layer module (workflow-agnostic)
- No imports from workflows or CLI
"""

import csv
import json
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime
from primerlab.core.models import WorkflowResult
from primerlab.core.logger import get_logger

logger = get_logger()


def generate_batch_summary(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Generates consolidated summary for batch runs.
    
    Args:
        results: List of result dictionaries from multiple workflow runs
        
    Returns:
        Dict containing:
        - total_sequences: int
        - successful: int
        - failed: int
        - summary_table: List of per-sequence stats
        - avg_quality_score: float
        - timestamp: str
    """
    total = len(results)
    successful = 0
    failed = 0
    summary_table = []
    quality_scores = []

    for i, result in enumerate(results):
        sequence_name = result.get("metadata", {}).get("sequence_name", f"Sequence_{i+1}")

        # Check if result has primers
        primers = result.get("primers", {})
        qc = result.get("qc", {})
        amplicons = result.get("amplicons", [])

        if primers:
            successful += 1
            status = "✅ Success"

            # Extract metrics
            fwd = primers.get("forward", {})
            rev = primers.get("reverse", {})

            fwd_tm = fwd.get("tm", 0)
            rev_tm = rev.get("tm", 0)
            product_size = amplicons[0].get("length", 0) if amplicons else 0
            quality_score = qc.get("quality_score", 0)

            if quality_score:
                quality_scores.append(quality_score)

            summary_table.append({
                "name": sequence_name,
                "status": status,
                "fwd_primer": fwd.get("id", "N/A"),
                "rev_primer": rev.get("id", "N/A"),
                "fwd_tm": round(fwd_tm, 1) if fwd_tm else "N/A",
                "rev_tm": round(rev_tm, 1) if rev_tm else "N/A",
                "product_size": product_size,
                "quality_score": quality_score or "N/A"
            })
        else:
            failed += 1
            status = "❌ Failed"
            error_msg = result.get("error", "No primers found")

            summary_table.append({
                "name": sequence_name,
                "status": status,
                "fwd_primer": "-",
                "rev_primer": "-",
                "fwd_tm": "-",
                "rev_tm": "-",
                "product_size": "-",
                "quality_score": "-",
                "error": error_msg
            })

    avg_quality = sum(quality_scores) / len(quality_scores) if quality_scores else 0

    return {
        "total_sequences": total,
        "successful": successful,
        "failed": failed,
        "success_rate": round(successful / total * 100, 1) if total > 0 else 0,
        "avg_quality_score": round(avg_quality, 1),
        "summary_table": summary_table,
        "timestamp": datetime.now().isoformat()
    }


def save_batch_summary_csv(summary: Dict[str, Any], filepath: str):
    """
    Saves batch summary to CSV file.
    
    Args:
        summary: Output from generate_batch_summary()
        filepath: Output file path
    """
    path = Path(filepath)

    try:
        with open(path, 'w', newline='') as f:
            writer = csv.writer(f)

            # Header
            writer.writerow([
                "Sequence", "Status", "Forward Primer", "Reverse Primer",
                "Fwd Tm", "Rev Tm", "Product Size", "Quality Score"
            ])

            # Data rows
            for row in summary["summary_table"]:
                writer.writerow([
                    row["name"],
                    row["status"],
                    row["fwd_primer"],
                    row["rev_primer"],
                    row["fwd_tm"],
                    row["rev_tm"],
                    row["product_size"],
                    row["quality_score"]
                ])

            # Summary footer
            writer.writerow([])
            writer.writerow(["SUMMARY"])
            writer.writerow(["Total Sequences", summary["total_sequences"]])
            writer.writerow(["Successful", summary["successful"]])
            writer.writerow(["Failed", summary["failed"]])
            writer.writerow(["Success Rate", f"{summary['success_rate']}%"])
            writer.writerow(["Avg Quality Score", summary["avg_quality_score"]])

        logger.info(f"Batch summary CSV saved to: {path}")
    except Exception as e:
        logger.error(f"Failed to save batch summary CSV: {e}")


def save_batch_excel(results: List[Dict[str, Any]], filepath: str):
    """
    Combines all batch results into single Excel workbook.
    
    Sheets:
    - Summary: Overview of all sequences
    - Per-sequence sheets with detailed info
    
    v0.1.5: New batch export
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed. Run: pip install openpyxl")
        return

    path = Path(filepath)
    summary = generate_batch_summary(results)

    try:
        wb = Workbook()

        # === SUMMARY SHEET ===
        ws_summary = wb.active
        ws_summary.title = "Batch Summary"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Summary stats
        stats = [
            ("Batch Summary Report", ""),
            ("", ""),
            ("Total Sequences", summary["total_sequences"]),
            ("Successful", summary["successful"]),
            ("Failed", summary["failed"]),
            ("Success Rate", f"{summary['success_rate']}%"),
            ("Avg Quality Score", summary["avg_quality_score"]),
            ("Generated", summary["timestamp"]),
        ]

        for row_idx, (label, value) in enumerate(stats, 1):
            cell_label = ws_summary.cell(row=row_idx, column=1, value=label)
            cell_value = ws_summary.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:
                cell_label.font = Font(bold=True, size=14)

        # Results table starting at row 10
        headers = ["Sequence", "Status", "Fwd Primer", "Rev Primer", 
                  "Fwd Tm", "Rev Tm", "Product Size", "Quality"]

        row_offset = 10
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=row_offset, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for i, row in enumerate(summary["summary_table"], row_offset + 1):
            ws_summary.cell(row=i, column=1, value=row["name"]).border = thin_border

            status_cell = ws_summary.cell(row=i, column=2, value=row["status"])
            status_cell.border = thin_border
            if "Success" in row["status"]:
                status_cell.fill = success_fill
            else:
                status_cell.fill = fail_fill

            ws_summary.cell(row=i, column=3, value=row["fwd_primer"]).border = thin_border
            ws_summary.cell(row=i, column=4, value=row["rev_primer"]).border = thin_border
            ws_summary.cell(row=i, column=5, value=row["fwd_tm"]).border = thin_border
            ws_summary.cell(row=i, column=6, value=row["rev_tm"]).border = thin_border
            ws_summary.cell(row=i, column=7, value=row["product_size"]).border = thin_border
            ws_summary.cell(row=i, column=8, value=row["quality_score"]).border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws_summary.column_dimensions[get_column_letter(col)].width = 15

        wb.save(path)
        logger.info(f"Batch Excel saved to: {path}")

    except Exception as e:
        logger.error(f"Failed to save batch Excel: {e}")


def format_batch_summary_cli(summary: Dict[str, Any]) -> str:
    """
    Format batch summary for CLI output.
    
    Args:
        summary: Output from generate_batch_summary()
        
    Returns:
        Formatted string for terminal display
    """
    lines = []

    lines.append("")
    lines.append("=" * 70)
    lines.append("📊 BATCH RUN SUMMARY")
    lines.append("=" * 70)
    lines.append("")
    lines.append(f"Total Sequences: {summary['total_sequences']}")
    lines.append(f"Successful: {summary['successful']}")
    lines.append(f"Failed: {summary['failed']}")
    lines.append(f"Success Rate: {summary['success_rate']}%")
    lines.append(f"Avg Quality Score: {summary['avg_quality_score']}")
    lines.append("")
    lines.append("-" * 70)
    lines.append("")

    # Table header
    lines.append(f"{'Sequence':<20} {'Status':<12} {'Fwd Tm':>8} {'Rev Tm':>8} {'Size':>8} {'QS':>6}")
    lines.append("-" * 70)

    for row in summary["summary_table"]:
        lines.append(
            f"{row['name']:<20} {row['status']:<12} "
            f"{str(row['fwd_tm']):>8} {str(row['rev_tm']):>8} "
            f"{str(row['product_size']):>8} {str(row['quality_score']):>6}"
        )

    lines.append("")
    lines.append("=" * 70)

    return "\n".join(lines)
